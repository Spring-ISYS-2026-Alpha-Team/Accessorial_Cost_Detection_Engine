"""
PACE — Port & Intermodal Terminal Dwell Time Collector
========================================================
Three sources:

  1. PMSA Dwell Time Reports — https://www.pmsaship.com/dwell-times-main
     Monthly San Pedro Bay (LA/Long Beach) truck and rail container dwell
     times scraped from plain-HTML Squarespace article pages.
     No key required.  requests + BeautifulSoup.

  2. BTS Port Performance — https://data.bts.gov/resource/69qe-yiui.json
     Annual port performance statistics (vessel calls, TEUs, berth length,
     cranes) for the top 25 US container ports via the Socrata SODA API.
     Also pulls the monthly TEU sub-dataset.
     No key required.

  3. STB Rail Service Data — https://www.stb.gov/reports-data/rail-service-data/
     Weekly Class I railroad consolidated Excel workbook.
     Contains terminal dwell times by railroad and individual terminal
     (Item 2), system train speeds (Item 1), and cars online (Item 3).
     Playwright headless Chromium is used to extract the current
     download link (it changes with each update), then openpyxl
     parses the multi-sheet workbook.

Outputs (./dwell_data/):
  pmsa_dwell_timeseries.csv        PMSA monthly truck + rail dwell 2019–present
  pmsa_scrape_log.csv              Per-URL status
  bts_port_performance.csv         Annual port capacity/throughput metrics
  bts_monthly_teus.csv             Monthly TEU counts top-10 ports
  stb_terminal_dwell.csv           Weekly average terminal dwell by railroad
  stb_train_speeds.csv             Weekly system average train speeds
  stb_cars_online.csv              Weekly cars online by car type
  dwell_pull_report.csv            Run summary

Setup:
  pip install requests beautifulsoup4 pandas openpyxl playwright
  playwright install chromium
"""

import re
import os
import io
import sys
import time
import tempfile
from datetime import datetime, date, timezone

# Force UTF-8 output on Windows
if sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl

OUT = "./dwell_data"
os.makedirs(OUT, exist_ok=True)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    )
}


# ═══════════════════════════════════════════════════════════════════════════
# PART 1 — PMSA DWELL TIMES
# ═══════════════════════════════════════════════════════════════════════════
#
# The PMSA Squarespace site publishes one article per month.
# Each article contains plain-prose sentences like:
#   "truck cargo...averaged 2.75 days"
#   "rail-destined containers averaged 6.14 days"
#
# Strategy:
#   Step A — scrape the index page to discover all article URLs
#             tagged under the "Dwell Times" category
#   Step B — fetch each article and regex-extract the two dwell values
#
# URL pattern for category filter:
#   The Squarespace JSON API uses a Unix-millisecond timestamp as the offset.
#   offset=N means "return up to 20 items published BEFORE timestamp N".
#   To start from the newest articles, use the current time (or a far-future ts).
#   Pagination: after each page, use the publishOn of the oldest item as the
#   next offset so the next call returns the next 20 older items.

PMSA_BLOG_BASE = "https://www.pmsaship.com"
PMSA_CATEGORY_API = (
    "https://www.pmsaship.com/maritime-insights-blog"
    "?category=Dwell+Times&format=json&offset={offset}"
)


def pmsa_discover_articles():
    """
    Hit the Squarespace JSON API to enumerate all dwell-time article URLs.
    Squarespace returns up to 20 items per page sorted newest-first.
    The offset parameter is a Unix millisecond timestamp — items published
    BEFORE that timestamp are returned.  Start from the current time and
    walk backwards using the publishOn of the last item on each page.
    Returns list of dicts: {title, url, published_date}
    """
    articles = []
    # Start just ahead of now so the first call returns the newest 20 articles
    offset = int(time.time() * 1000) + 1
    page   = 1

    while True:
        url = PMSA_CATEGORY_API.format(offset=offset)
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"    PMSA discovery error at offset {offset}: {e}")
            break

        items = data.get("items", [])
        if not items:
            break

        for item in items:
            full_url = PMSA_BLOG_BASE + item.get("fullUrl", "")
            title    = item.get("title", "")
            pub_ts   = item.get("publishOn", item.get("addedOn", 0))
            pub_date = datetime.fromtimestamp(pub_ts / 1000).strftime("%Y-%m-%d") if pub_ts else ""
            articles.append({"title": title, "url": full_url, "published_date": pub_date})

        print(f"    Page {page}: +{len(items)} articles  (total so far: {len(articles)})")

        # Squarespace: pagination is done — no more pages when fewer than 20 items
        if len(items) < 20:
            break

        # Next offset = publishOn of the oldest item on this page (last in list)
        last_pub_ts = items[-1].get("publishOn", items[-1].get("addedOn", 0))
        if not last_pub_ts or last_pub_ts >= offset:
            break  # guard against infinite loop
        offset = last_pub_ts
        page   += 1
        time.sleep(0.5)

    return articles


# Regex patterns for extracting dwell values from article body text.
# PMSA uses several phrasings across different report years:
#   "truck cargo...averaged 2.75 days"
#   "local truck dwell averaged 2.73 days"
#   "truck dwell times averaging 2.87 days"
#   "rail dwell times at 3.98 days"
#   "rail-destined containers averaged 6.14 days"

TRUCK_PATTERNS = [
    r"(?:local\s+)?truck(?:\s+cargo)?(?:\s+dwell)?(?:\s+time[s]?)?\s+(?:averaged?|averaging|at|of|remain(?:ed)?(?:\s+at)?|came\s+in\s+at)\s+([\d]+\.[\d]+)\s+days",
    r"truck(?:\s+dwell)?\s+averaged?\s+([\d]+\.[\d]+)\s+days",
    r"truck\s+dwell\s+(?:times?\s+)?(?:at|of|averaging)\s+([\d]+\.[\d]+)\s+days",
    r"([\d]+\.[\d]+)\s+days?\s+(?:for|at)\s+(?:local\s+)?truck",
    # "via truck spent an average of 2.75 days"  /  "truck cargo...average of 2.87 days"
    r"truck[^.]{0,120}?(?:an\s+)?average\s+of\s+([\d]+\.[\d]+)\s+days",
]

RAIL_PATTERNS = [
    r"rail(?:[- ]destined)?(?:\s+containers?)?(?:\s+dwell)?(?:\s+time[s]?)?\s+(?:averaged?|averaging|at|of|remain(?:ed)?(?:\s+at)?|came\s+in\s+at)\s+([\d]+\.[\d]+)\s+days",
    r"rail\s+dwell\s+(?:times?\s+)?(?:at|of|averaging|improved\s+to|rose\s+to)\s+([\d]+\.[\d]+)\s+days",
    r"rail\s+averaged?\s+([\d]+\.[\d]+)\s+days",
    r"([\d]+\.[\d]+)\s+days?\s+(?:for|at)\s+rail",
]

# Month/year extraction from article title or URL
MONTH_PATTERN = re.compile(
    r"(january|february|march|april|may|june|july|august|september|october|november|december)"
    r"[\s\-]+(\d{4})",
    re.IGNORECASE
)
MONTH_MAP = {
    "january":1,"february":2,"march":3,"april":4,"may":5,"june":6,
    "july":7,"august":8,"september":9,"october":10,"november":11,"december":12
}


def _extract_dwell(text, patterns):
    """Return first float matched by any pattern in [0, 30] range."""
    for pat in patterns:
        for m in re.finditer(pat, text, re.IGNORECASE):
            try:
                val = float(m.group(1))
                if 0.0 < val < 30.0:
                    return val
            except (ValueError, IndexError):
                continue
    return None


def pmsa_scrape_article(url, title, published_date):
    """
    Fetch one PMSA article and extract truck_dwell_days + rail_dwell_days.
    Returns a dict.
    """
    row = {
        "published_date":  published_date,
        "report_month":    None,
        "report_year":     None,
        "title":           title,
        "url":             url,
        "truck_dwell_days": None,
        "rail_dwell_days":  None,
        "scrape_status":   "ok",
    }

    # Parse report period from title/url
    text_for_date = (title or "") + " " + (url or "")
    dm = MONTH_PATTERN.search(text_for_date)
    if dm:
        row["report_month"] = MONTH_MAP[dm.group(1).lower()]
        row["report_year"]  = int(dm.group(2))

    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code == 404:
            row["scrape_status"] = "404"
            return row
        r.raise_for_status()
    except requests.RequestException as e:
        row["scrape_status"] = f"error:{e}"
        return row

    soup = BeautifulSoup(r.text, "html.parser")
    # Squarespace body content is in <div class="sqs-block-content"> blocks
    content_blocks = soup.find_all("div", class_=re.compile(r"sqs-block-content|entry-content|blog-item-content"))
    if content_blocks:
        body_text = " ".join(b.get_text(separator=" ") for b in content_blocks)
    else:
        body_text = soup.get_text(separator=" ")
    body_text = re.sub(r"\s+", " ", body_text)

    row["truck_dwell_days"] = _extract_dwell(body_text, TRUCK_PATTERNS)
    row["rail_dwell_days"]  = _extract_dwell(body_text, RAIL_PATTERNS)

    if row["truck_dwell_days"] is None and row["rail_dwell_days"] is None:
        row["scrape_status"] = "no_values_found"

    return row


def pull_pmsa():
    print("\n" + "=" * 65)
    print("  PART 1 — PMSA Dwell Time Reports")
    print("  Source: pmsaship.com (Squarespace JSON API + article scrape)")
    print("=" * 65)

    # Step A: discover all dwell-time articles
    print("\n  Discovering article URLs...")
    articles = pmsa_discover_articles()
    print(f"  Found {len(articles)} dwell-time articles\n")

    if not articles:
        print("  No articles found — check PMSA site structure")
        return {}

    # Step B: scrape each article
    rows     = []
    log_rows = []
    ok = no_val = err = 0

    for i, art in enumerate(articles, 1):
        row = pmsa_scrape_article(art["url"], art["title"], art["published_date"])

        status   = row["scrape_status"]
        truck    = row["truck_dwell_days"]
        rail     = row["rail_dwell_days"]
        period   = f"{row['report_year']}-{row['report_month']:02d}" if row["report_year"] else "?"

        if status == "ok":
            sym = "✓"; ok += 1
        elif status == "no_values_found":
            sym = "~"; no_val += 1
        else:
            sym = "✗"; err += 1

        detail = ""
        if truck: detail += f"  truck={truck:.2f}d"
        if rail:  detail += f"  rail={rail:.2f}d"

        print(f"  [{i:03d}/{len(articles)}]  {period:<8}  {sym}  {art['title'][:45]:<46}{detail}")

        rows.append(row)
        log_rows.append({
            "period":       period,
            "title":        art["title"],
            "url":          art["url"],
            "status":       status,
            "truck_dwell":  truck,
            "rail_dwell":   rail,
        })
        time.sleep(0.6)

    # Save results
    df = (pd.DataFrame(rows)
            .sort_values(["report_year","report_month"])
            .reset_index(drop=True))

    ts_path  = os.path.join(OUT, "pmsa_dwell_timeseries.csv")
    log_path = os.path.join(OUT, "pmsa_scrape_log.csv")
    df.to_csv(ts_path, index=False)
    pd.DataFrame(log_rows).to_csv(log_path, index=False)

    print(f"\n  ✓  {ts_path}  ({len(df)} rows)")
    print(f"  ✓  {log_path}")
    print(f"\n  ok={ok}  no_values={no_val}  error={err}")

    # Show recent readings
    preview = df.dropna(subset=["truck_dwell_days"]).tail(8)[
        ["report_year","report_month","truck_dwell_days","rail_dwell_days"]
    ]
    print(f"\n  Last 8 readings with truck dwell:")
    print(preview.to_string(index=False))

    return {"pmsa_rows": len(df), "pmsa_ok": ok, "pmsa_err": err}


# ═══════════════════════════════════════════════════════════════════════════
# PART 2 — BTS PORT PERFORMANCE (Socrata SODA API)
# ═══════════════════════════════════════════════════════════════════════════
#
# BTS hosts several port datasets on data.bts.gov via the Socrata platform.
# All use the same SODA REST API pattern:
#   GET https://data.bts.gov/resource/{dataset_id}.json
#       ?$limit=5000&$offset=0&$order=:id
#
# Datasets pulled:
#   69qe-yiui  — Port Performance Freight Statistics (annual, top 25 ports)
#                Columns: port_name, year, vessel_calls, teus, berth_length_ft,
#                         container_cranes, on_dock_rail, pti_lalb, pti_nynj
#
#   rd72-aq8r  — Monthly TEU Data (monthly, top 10 ports)
#                Columns: port, month, year, inbound_full, outbound_full,
#                         inbound_empty, outbound_empty, total_teus
#
# Max 5,000 rows per call; paginate with $offset.

BTS_DATASETS = [
    # NOTE: the annual port performance dataset (69qe-yiui) was removed by BTS as of 2025.
    # Only the monthly TEU dataset remains publicly accessible via the SODA API.
    {
        "id":    "rd72-aq8r",
        "name":  "Monthly TEU Data",
        "file":  "bts_monthly_teus.csv",
        "desc":  "Monthly TEU counts at top-10 US container ports (2019–present)",
    },
]

SODA_BASE = "https://data.bts.gov/resource"
SODA_LIMIT = 5000


def socrata_fetch_all(dataset_id, dataset_name):
    """
    Pull all rows from a BTS Socrata dataset using $limit/$offset pagination.
    Returns a DataFrame.
    """
    all_rows = []
    offset   = 0
    page     = 1

    while True:
        url = f"{SODA_BASE}/{dataset_id}.json"
        params = {
            "$limit":  SODA_LIMIT,
            "$offset": offset,
            "$order":  ":id",
        }
        try:
            r = requests.get(url, params=params, headers=HEADERS, timeout=30)
            r.raise_for_status()
            rows = r.json()
        except Exception as e:
            print(f"    SODA error page {page}: {e}")
            break

        if not rows:
            break

        all_rows.extend(rows)
        print(f"    page {page}: +{len(rows)} rows  (total: {len(all_rows)})", end="\r")

        if len(rows) < SODA_LIMIT:
            break

        offset += SODA_LIMIT
        page   += 1
        time.sleep(0.3)

    print()  # newline after carriage-return
    return pd.DataFrame(all_rows) if all_rows else pd.DataFrame()


def pull_bts():
    print("\n" + "=" * 65)
    print("  PART 2 — BTS Port Performance (Socrata SODA API)")
    print("  Base: data.bts.gov/resource/{id}.json")
    print("=" * 65)

    results = {}

    for ds in BTS_DATASETS:
        print(f"\n  [{ds['id']}]  {ds['name']}")
        print(f"  {ds['desc']}")

        df = socrata_fetch_all(ds["id"], ds["name"])

        if df.empty:
            print(f"  ✗  No data returned")
            results[ds["id"]] = 0
            continue

        # Clean column names: lowercase, replace spaces/special chars with underscore
        df.columns = [
            re.sub(r"[^a-z0-9_]", "_", c.lower().strip())
            for c in df.columns
        ]

        path = os.path.join(OUT, ds["file"])
        df.to_csv(path, index=False)
        print(f"  ✓  {len(df):,} rows × {len(df.columns)} cols  →  {path}")
        print(f"     Columns: {list(df.columns[:10])}{'...' if len(df.columns)>10 else ''}")
        results[ds["id"]] = len(df)
        time.sleep(0.5)

    return results


# ═══════════════════════════════════════════════════════════════════════════
# PART 3 — STB RAIL SERVICE DATA (Playwright + openpyxl)
# ═══════════════════════════════════════════════════════════════════════════
#
# The STB page at stb.gov/reports-data/rail-service-data/ lists the
# consolidated "All Class I Railroads" Excel workbook as a dynamically
# rendered link. The download URL follows this pattern:
#   https://dcms-external.s3.amazonaws.com/DCMS_External_PROD/{id}/...xlsx
#
# Strategy:
#   1. Use Playwright to render the STB page and extract the .xlsx link
#   2. Download the file with requests
#   3. Parse with openpyxl — the workbook has named sheets for each item:
#        Item 1  = System average train speed by train type
#        Item 2  = Weekly average terminal dwell time (by railroad + terminal)
#        Item 3  = Cars online by car type
#        (Items 4–9 are grain/coal/carload data, not needed for PACE)
#
# Sheet structure (confirmed from STB documentation and public analysis):
#   Item 2 has header rows followed by data rows in a wide format where:
#     - Column A = railroad name
#     - Column B = metric label (System, or terminal name)
#     - Remaining columns = weekly report dates (one date per column)
#   We transpose this to long format: railroad, terminal, report_date, dwell_hours

STB_PAGE_URL = "https://www.stb.gov/reports-data/rail-service-data/"

# As of 2026 the file moved from S3 to WordPress hosting on stb.gov.
# Both patterns are checked so the scraper survives future moves.
STB_XLSX_PATTERN = re.compile(
    r"https://(?:"
    r"www\.stb\.gov/wp-content/uploads/[^\s\"']+\.xlsx"
    r"|dcms-external\.s3\.amazonaws\.com/[^\s\"']+\.xlsx"
    r")",
    re.IGNORECASE
)

# Fallback URL — most recent consolidated file as of 2026-03-11
STB_FALLBACK_URL = (
    "https://www.stb.gov/wp-content/uploads/files/rsir/"
    "All%20Class%201%20Railroads/"
    "EP724%20Consolidated%20Data%20through%202026-03-11.xlsx"
)

# Class I railroads that report to STB
CLASS_I_RAILROADS = ["BNSF", "CN", "CPKC", "CSX", "NS", "UP"]


def stb_get_xlsx_url():
    """
    Use Playwright to render the STB page and find the consolidated
    Excel download link. Falls back to Playwright page source scrape
    if link isn't in plain HTML.
    Returns download URL string or None.
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  Playwright not installed — using fallback URL")
        return STB_FALLBACK_URL

    print("  Launching Playwright to find STB Excel URL...")

    found_url = None
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        page    = browser.new_page(user_agent=HEADERS["User-Agent"])

        try:
            page.goto(STB_PAGE_URL, wait_until="networkidle", timeout=30_000)
            page.wait_for_timeout(2000)

            # Search all anchor hrefs for an .xlsx link (WordPress or S3)
            links = page.eval_on_selector_all(
                "a[href]",
                "els => els.map(e => e.href)"
            )
            for link in links:
                if STB_XLSX_PATTERN.search(link) and (
                    "All_Class" in link or "Consolidated" in link or "EP724" in link
                ):
                    found_url = link
                    break

            # Broader fallback: any matching .xlsx link
            if not found_url:
                for link in links:
                    if STB_XLSX_PATTERN.search(link):
                        found_url = link
                        break

            # Last resort: search full page HTML
            if not found_url:
                html_content = page.content()
                m = STB_XLSX_PATTERN.search(html_content)
                if m:
                    found_url = m.group(0)

        except Exception as e:
            print(f"  Playwright error: {e}")
        finally:
            browser.close()

    if found_url:
        print(f"  Found STB Excel URL: {found_url[:80]}...")
    else:
        print(f"  Could not find URL via Playwright — using fallback")
        found_url = STB_FALLBACK_URL

    return found_url


def stb_download_xlsx(url):
    """Download the STB Excel file. Returns bytes or None."""
    print(f"  Downloading STB Excel workbook...")
    try:
        r = requests.get(url, headers=HEADERS, timeout=60, stream=True)
        r.raise_for_status()
        content = b"".join(r.iter_content(chunk_size=8192))
        size_mb = len(content) / (1024 * 1024)
        print(f"  Downloaded {size_mb:.1f} MB")
        return content
    except requests.RequestException as e:
        print(f"  Download failed: {e}")
        return None


def _is_date_like(val):
    """Check if a cell value looks like a report date."""
    if isinstance(val, datetime):
        return True
    if isinstance(val, str):
        return bool(re.match(r"\d{1,2}/\d{1,2}/\d{2,4}", val.strip()))
    return False


def _to_date_str(val):
    """Normalize a date value to YYYY-MM-DD string."""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        try:
            return pd.to_datetime(val, errors="coerce").strftime("%Y-%m-%d")
        except Exception:
            return str(val)
    return str(val)


def _stb_load_tidy(wb):
    """
    Load the STB workbook into a tidy long-format DataFrame.

    New format (as of 2026): single Sheet1 with columns:
      A: Railroad/Region
      B: Category No.  (1=Train Speed, 2=Terminal Dwell, 3=Cars Online, ...)
      C: Sub-Category
      D: Measure
      E: Variable       (terminal name / train type / car type)
      F: Sub-Variable
      G+: Weekly date columns (datetime values in header row 1)

    Returns (df_wide, date_cols) where date_cols is a list of column names
    that are date strings, or (None, []) on failure.
    """
    ws = wb.worksheets[0]
    print(f"    Loading sheet: '{ws.title}'  ({ws.max_row} rows x {ws.max_column} cols)")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, []

    header = rows[0]
    # Find first date column in header
    date_col_start = None
    for ci, v in enumerate(header):
        if _is_date_like(v):
            date_col_start = ci
            break

    if date_col_start is None:
        print("    Could not locate date columns in header row")
        return None, []

    # Build column names: fixed names for A-F, date strings for G+
    fixed_names = ["railroad", "category_no", "sub_category", "measure", "variable", "sub_variable"]
    col_names = fixed_names[:date_col_start]
    date_cols = []
    for ci in range(date_col_start, len(header)):
        v = header[ci]
        if v is not None:
            ds = _to_date_str(v)
            col_names.append(ds)
            date_cols.append(ds)
        else:
            col_names.append(f"_col{ci}")

    df = pd.DataFrame(rows[1:], columns=col_names[:len(rows[1])] if rows[1:] else col_names)
    df["railroad"]    = df["railroad"].ffill()   # railroad repeats in col A across its rows
    df["category_no"] = df["category_no"].astype(str).str.strip()

    print(f"    {len(df):,} data rows  |  {len(date_cols)} weekly dates  "
          f"({date_cols[0]} -> {date_cols[-1]})")

    return df, date_cols


def stb_parse_item2_dwell(wb):
    """
    Parse Item 2 (Weekly Average Terminal Dwell Time) from the STB workbook.
    New format: single tidy Sheet1, filtered by Category No. == '2'.
    Returns DataFrame: railroad, terminal, report_date, avg_terminal_dwell_hours
    """
    df, date_cols = _stb_load_tidy(wb)
    if df is None or not date_cols:
        return pd.DataFrame()

    sub = df[df["category_no"] == "2"].copy()
    if sub.empty:
        print("    No Category 2 (Terminal Dwell) rows found")
        return pd.DataFrame()

    records = []
    for _, row in sub.iterrows():
        terminal = str(row.get("variable", "")).strip() or "System"
        railroad = str(row.get("railroad", "")).strip()
        for dc in date_cols:
            val = row.get(dc)
            if val is None:
                continue
            try:
                dwell_hours = float(val)
                if not (0 < dwell_hours < 500):
                    continue
            except (ValueError, TypeError):
                continue
            records.append({
                "railroad":                 railroad,
                "terminal":                 terminal,
                "report_date":              dc,
                "avg_terminal_dwell_hours": dwell_hours,
            })

    result = pd.DataFrame(records)
    if not result.empty:
        result["report_date"] = pd.to_datetime(result["report_date"], errors="coerce")
        result = result.dropna(subset=["report_date"])
        result["report_date"] = result["report_date"].dt.strftime("%Y-%m-%d")
        result = result.sort_values(["railroad", "terminal", "report_date"]).reset_index(drop=True)
    return result


def stb_parse_item1_speeds(wb):
    """
    Parse Item 1 (System Average Train Speed) from the STB workbook.
    New format: single tidy Sheet1, filtered by Category No. == '1'.
    Returns DataFrame: railroad, train_type, report_date, avg_speed_mph
    """
    df, date_cols = _stb_load_tidy(wb)
    if df is None or not date_cols:
        return pd.DataFrame()

    sub = df[df["category_no"] == "1"].copy()
    if sub.empty:
        print("    No Category 1 (Train Speed) rows found")
        return pd.DataFrame()

    records = []
    for _, row in sub.iterrows():
        train_type = str(row.get("variable", "")).strip() or "System"
        railroad   = str(row.get("railroad", "")).strip()
        for dc in date_cols:
            val = row.get(dc)
            if val is None:
                continue
            try:
                speed = float(val)
                if not (0 < speed < 120):
                    continue
            except (ValueError, TypeError):
                continue
            records.append({
                "railroad":      railroad,
                "train_type":    train_type,
                "report_date":   dc,
                "avg_speed_mph": speed,
            })

    result = pd.DataFrame(records)
    if not result.empty:
        result["report_date"] = pd.to_datetime(result["report_date"], errors="coerce")
        result = result.dropna(subset=["report_date"])
        result["report_date"] = result["report_date"].dt.strftime("%Y-%m-%d")
        result = result.sort_values(["railroad", "train_type", "report_date"]).reset_index(drop=True)
    return result


def stb_parse_item3_cars(wb):
    """
    Parse Item 3 (Cars Online by Car Type) from the STB workbook.
    New format: single tidy Sheet1, filtered by Category No. == '3'.
    Returns DataFrame: railroad, car_type, report_date, cars_online
    """
    df, date_cols = _stb_load_tidy(wb)
    if df is None or not date_cols:
        return pd.DataFrame()

    sub = df[df["category_no"] == "3"].copy()
    if sub.empty:
        print("    No Category 3 (Cars Online) rows found")
        return pd.DataFrame()

    records = []
    for _, row in sub.iterrows():
        car_type = str(row.get("variable", "")).strip() or "Total"
        railroad = str(row.get("railroad", "")).strip()
        for dc in date_cols:
            val = row.get(dc)
            if val is None:
                continue
            try:
                cars = float(val)
                if not (0 < cars < 2_000_000):
                    continue
            except (ValueError, TypeError):
                continue
            records.append({
                "railroad":    railroad,
                "car_type":    car_type,
                "report_date": dc,
                "cars_online": int(cars),
            })

    result = pd.DataFrame(records)
    if not result.empty:
        result["report_date"] = pd.to_datetime(result["report_date"], errors="coerce")
        result = result.dropna(subset=["report_date"])
        result["report_date"] = result["report_date"].dt.strftime("%Y-%m-%d")
        result = result.sort_values(["railroad", "car_type", "report_date"]).reset_index(drop=True)

    return df


def pull_stb():
    print("\n" + "=" * 65)
    print("  PART 3 — STB Rail Service Data (Playwright + openpyxl)")
    print("  Source: stb.gov/reports-data/rail-service-data/")
    print("=" * 65)

    # Step 1: Get the current Excel download URL
    xlsx_url = stb_get_xlsx_url()
    if not xlsx_url:
        print("  Could not determine STB Excel URL — skipping")
        return {}

    # Step 2: Download the workbook
    xlsx_bytes = stb_download_xlsx(xlsx_url)
    if not xlsx_bytes:
        return {}

    # Save raw Excel for reference
    raw_path = os.path.join(OUT, "stb_consolidated_raw.xlsx")
    with open(raw_path, "wb") as f:
        f.write(xlsx_bytes)
    print(f"  Raw Excel saved: {raw_path}")

    # Step 3: Open with openpyxl and list sheets
    wb = openpyxl.load_workbook(
        io.BytesIO(xlsx_bytes),
        read_only=True,
        data_only=True
    )
    print(f"\n  Workbook sheets ({len(wb.sheetnames)}): {wb.sheetnames}")

    results = {}

    # Parse Item 2: Terminal Dwell (most PACE-relevant)
    print("\n  Parsing Item 2 — Terminal Dwell...")
    df_dwell = stb_parse_item2_dwell(wb)
    if not df_dwell.empty:
        path = os.path.join(OUT, "stb_terminal_dwell.csv")
        df_dwell.to_csv(path, index=False)
        print(f"  ✓  {len(df_dwell):,} rows → {path}")

        # Print summary: latest week by railroad
        latest = (df_dwell[df_dwell["terminal"] == "System"]
                  .sort_values("report_date")
                  .groupby("railroad")
                  .last()
                  .reset_index()[["railroad","report_date","avg_terminal_dwell_hours"]])
        if not latest.empty:
            print(f"\n  Latest system dwell by railroad:")
            print(latest.to_string(index=False))

        results["stb_dwell_rows"] = len(df_dwell)
    else:
        print("  ✗  No dwell data extracted")
        results["stb_dwell_rows"] = 0

    # Parse Item 1: Train Speeds
    print("\n  Parsing Item 1 — Train Speeds...")
    df_speed = stb_parse_item1_speeds(wb)
    if not df_speed.empty:
        path = os.path.join(OUT, "stb_train_speeds.csv")
        df_speed.to_csv(path, index=False)
        print(f"  ✓  {len(df_speed):,} rows → {path}")
        results["stb_speed_rows"] = len(df_speed)
    else:
        print("  ✗  No speed data extracted")
        results["stb_speed_rows"] = 0

    # Parse Item 3: Cars Online
    print("\n  Parsing Item 3 — Cars Online...")
    df_cars = stb_parse_item3_cars(wb)
    if not df_cars.empty:
        path = os.path.join(OUT, "stb_cars_online.csv")
        df_cars.to_csv(path, index=False)
        print(f"  ✓  {len(df_cars):,} rows → {path}")
        results["stb_cars_rows"] = len(df_cars)
    else:
        print("  ✗  No cars-online data extracted")
        results["stb_cars_rows"] = 0

    wb.close()
    return results


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    t0 = datetime.now(timezone.utc).replace(tzinfo=None)
    print("\n" + "#" * 65)
    print("  PACE — Port & Intermodal Dwell Time Collector")
    print(f"  {t0.strftime('%Y-%m-%d %H:%M UTC')}  |  {os.path.abspath(OUT)}")
    print("#" * 65)

    report = {"run_timestamp": t0.strftime("%Y-%m-%d %H:%M UTC")}

    pmsa_results = pull_pmsa()
    report.update(pmsa_results)

    bts_results = pull_bts()
    report.update({f"bts_{k}": v for k, v in bts_results.items()})

    stb_results = pull_stb()
    report.update(stb_results)

    elapsed = (datetime.now(timezone.utc).replace(tzinfo=None) - t0).total_seconds()
    report["elapsed_s"] = round(elapsed, 1)

    pd.DataFrame([report]).to_csv(os.path.join(OUT, "dwell_pull_report.csv"), index=False)

    print("\n" + "#" * 65)
    print(f"  Done in {elapsed:.0f}s")
    print(f"\n  Files in {os.path.abspath(OUT)}:")
    for f in sorted(os.listdir(OUT)):
        if f.endswith((".csv", ".xlsx")):
            kb = os.path.getsize(os.path.join(OUT, f)) / 1024
            print(f"    {f:<50}  {kb:>8.1f} KB")
    print("#" * 65 + "\n")


if __name__ == "__main__":
    main()
